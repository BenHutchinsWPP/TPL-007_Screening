from pathlib import Path
import pandas as pd
import openpyxl.utils
import win32com.client

mva_mismatch_threshold = 1.0 

def chk(SimAuto, SimAutoOutput, Message) -> list[list[str]]:
    """
    Function used to catch and display errors passed back from SimAuto

    SimAuto return object format:
    [0] = Error message, if any
    [1+] = Return data
    """

    if SimAutoOutput[0] != '':
        print('Error: ' + SimAutoOutput[0])
        return None
    # else:
    #     print(Message)

    if len(SimAutoOutput) == 1:
        return None
    elif len(SimAutoOutput) == 2:
        return SimAutoOutput[1]
    else:
        return SimAutoOutput[1:]

def get_param(SimAuto, table: str, parameters: list[str], filter_group: str = '') -> list[list[str]]:
    msg = 'GetParametersMultipleElementRect(' + table + ': [' + ', '.join(parameters) + '])'
    return_value: list[list[str]] = chk(SimAuto, SimAuto.GetParametersMultipleElementRect(table, parameters, filter_group), msg)
    return return_value

def get_table_types(SimAuto, table: str) -> dict[str,str]:
    columns = ['key', 'variablenamelegacy', 'type', 'description', 'concisename', 'enterable']
    fields_df = pd.DataFrame(SimAuto.GetFieldList(table)[1], columns=columns)

    type_map = {
        "String": "string",
        "Integer": "Int64",   # nullable integer
        "Real": "Float64"     # nullable float
    }

    table_types = {
        row["concisename"]: type_map[row["type"]]
        for _, row in fields_df.iterrows()
    }

    return table_types

def get_param_df(SimAuto, table: str, parameter_type: dict[str,type] | list[str], filter_group: str = '') -> pd.DataFrame:
    def deep_strip(obj):
        if isinstance(obj, str):
            return obj.strip()
        elif isinstance(obj, tuple):
            return tuple(deep_strip(x) for x in obj)
        elif isinstance(obj, list):
            return [deep_strip(x) for x in obj]
        elif isinstance(obj, dict):
            return {k: deep_strip(v) for k, v in obj.items()}
        elif isinstance(obj, set):
            return {deep_strip(x) for x in obj}
        else:
            return obj

    def apply_table_types(df, table_types):
        for col, dtype in table_types.items():
            if col not in df.columns:
                continue  # skip missing columns

            if dtype == "string":
                df[col] = df[col].astype("string")
            else:
                df[col] = pd.to_numeric(
                    df[col].replace("", pd.NA),
                    errors="coerce"
                ).astype(dtype)
        return df

    # Get data from PowerWorld. 
    if type(parameter_type) == dict:
        parameter_list: list[str] = list(parameter_type.keys())
    else:
        parameter_list: list[str] = parameter_type
    rows: list[list[str]] = get_param(SimAuto, table, parameter_list, filter_group)
    trimmed_rows = deep_strip(rows)

    # Pack into a dataframe. 
    table_types = get_table_types(SimAuto, table)
    df = pd.DataFrame(data=trimmed_rows, columns=parameter_list)
    df = apply_table_types(df, table_types)

    return df

def set_param(SimAuto, table: str, parameters: list[str], rows: list[list[str]]):
    msg = 'ChangeParametersMultipleElementRect(' + table + ': [' + ', '.join(parameters) + '])'
    return_value = chk(SimAuto, SimAuto.ChangeParametersMultipleElementRect(table, parameters, rows), msg)
    return return_value

def set_param_df(SimAuto, table: str, df: pd.DataFrame):
    df = df.reset_index()
    if(len(df) == 0):
        return ''

    # Get parameters. 
    parameters: list[str] = df.columns.tolist()
    # Convert df into list of lists. All numerical values which are "nan" must be treated as empty strings. 
    rows: list[list[str]] = df.astype("string").fillna("").values.tolist()
    # Set data in PowerWorld. 
    return_value = set_param(SimAuto, table, parameters, rows)
    return return_value

def open_case(SimAuto, fp) -> bool:
    # Attempts to open a case.
    # Error case: message = ('OpenCase: Errors have occurred',)
    # Success case: message = ('',)

    if not Path(fp).exists():
        print(f'Path does not exist: {str(fp)}')
        return False
    
    message = SimAuto.OpenCase(fp)

    if 'OpenCase: Error' in message[0]:
        print(f'Could not open: {str(fp)}')
        return False

    print(f'Opened: {str(fp)}')
    return True

def save_case(SimAuto, fp, case_format = 'PWB22') -> bool:
    # Attempts to save a case.
    # No case in memory to save out?
    #   Error case: message = ('SaveCase: Error trying to save c:\\case 1.pwb - aborted',)
    # Writing to a parent folder which doesn't exist?
    #   Error case: message = ('SaveCase: Windows has prevented us from writing to the file c:\\foobar\\nonexistent directory.pwb. Verify your write privileges and that the specified parent directory exists.',)
    # Success case: message = ('',)

    if not Path(fp).parent.exists():
        print(f'Path does not exist: {str(fp)}')
        return False
    
    message = SimAuto.SaveCase(fp, case_format, True)

    if 'SaveCase: ' in message[0]:
        print(f'Could not save to: {str(fp)}')
        print(message[0])
        return False

    print(f'Saved: {str(fp)}')
    return True

def solve(SimAuto, mva_mismatch_threshold = 1.0) -> bool:
    # Solve.
    SimAuto.RunScriptCommand('EnterMode(RUN);')
    result = SimAuto.RunScriptCommand('SolvePowerFlow(RECTNEWT);')

    # Error string. Return early with False if it didn't solve. 
    if result[0] != '': 
        print(result[0])
        return False
    SimAuto.RunScriptCommand('EnterMode(EDIT);')

    # Get mismatch. 
    df = get_param_df(SimAuto, 'Bus', {'Busnum':int, 'MismatchP':float, 'MismatchQ':float})
    df['MismatchS'] = (df['MismatchP']**2.0 + df['MismatchQ']**2.0)**0.5
    max_mismatch = df['MismatchS'].abs().max()

    # print(f'Max Mismatch (S) = {max_mismatch}')

    return max_mismatch < mva_mismatch_threshold

def auto_fit_columns(writer: pd.ExcelWriter):
    workbook = writer.book
    for sheet_name in writer.sheets:
        worksheet = workbook[sheet_name]
        for col in worksheet.columns:
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
                except:
                    pass
            worksheet.column_dimensions[col_letter].width = max_length + 2

def freeze_top_rows(writer: pd.ExcelWriter):
    for sheet in writer.sheets.values():
        sheet.freeze_panes = 'A2'  # Freeze the top row
    return

def filter_top_rows(writer: pd.ExcelWriter):
    for sheet in writer.sheets.values():
        max_column = sheet.max_column
        max_column_letter = openpyxl.utils.get_column_letter(max_column)
        sheet.auto_filter.ref = f"A1:{max_column_letter}1"  # Adjust range based on columns
    return

def df_dict_to_excel_workbook(rep_fp: Path, dict_df: dict[str,pd.DataFrame]):
    """Writes a dictionary of dataframes to an Excel Workbook."""
    writer = pd.ExcelWriter(rep_fp, engine='openpyxl')
    for sheet_name, df in dict_df.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    auto_fit_columns(writer)
    freeze_top_rows(writer)
    filter_top_rows(writer)
    try:
        writer.close()
    except:
        pass
